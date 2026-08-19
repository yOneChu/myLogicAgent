#!/usr/bin/env python3
"""
엘리베이터 특성코드 업데이트 검사 스크립트

master_data/엘리베이터_특성코드.xlsx 파일과 
최신 API(getCodeList)의 응답 데이터를 비교하여
신규 추가, 수정, 삭제된 특성코드가 있는지 검사합니다.
"""

import json
import ssl
import sys
import urllib.request
import urllib.parse
from pathlib import Path
from typing import Dict, List, Tuple, Any
from openpyxl import load_workbook

# API URL 및 접근 키 설정
API_URL = "https://vault-in.hdel.co.kr:8070/api/getCodeList"
API_KEY = "subae"

# 워크스페이스 최상위 경로 지정
WORKSPACE_ROOT = Path(__file__).resolve().parent.parent
MASTER_DATA_DIR = WORKSPACE_ROOT / "master_data"
EXCEL_PATH = MASTER_DATA_DIR / "엘리베이터_특성코드.xlsx"

def fetch_latest_code_list() -> List[Dict[str, str]]:
    """
    최신 엘리베이터 특성코드 API를 호출하여 최신 특성코드 리스트를 가져오는 함수.
    
    Returns:
        List[Dict[str, str]]: API 응답 특성코드 객체 리스트
    """
    query_string = urllib.parse.urlencode({"key": API_KEY})
    url = f"{API_URL}?{query_string}"
    
    req = urllib.request.Request(
        url,
        method="GET",
        headers={
            "Accept": "application/json",
            "User-Agent": "elevator-code-checker/1.0"
        }
    )
    
    # 사내 SSL 인증서 검증 생략
    ctx = ssl._create_unverified_context()
    
    with urllib.request.urlopen(req, timeout=120, context=ctx) as response:
        raw_bytes = response.read()
        
    try:
        text = raw_bytes.decode("utf-8")
    except UnicodeDecodeError:
        text = raw_bytes.decode("cp949", errors="replace")
        
    data = json.loads(text)
    
    if isinstance(data, list):
        return data
    elif isinstance(data, dict):
        for key in ("data", "result", "items", "rows", "list"):
            if key in data and isinstance(data[key], list):
                return data[key]
    
    raise ValueError("API 응답 데이터 형식을 해석할 수 없습니다.")

def read_existing_excel(excel_path: Path) -> List[Dict[str, str]]:
    """
    master_data/엘리베이터_특성코드.xlsx 파일에서 기존 특성코드 데이터를 읽어오는 함수.
    
    Args:
        excel_path (Path): 엑셀 파일 경로
        
    Returns:
        List[Dict[str, str]]: 기존 엑셀 데이터 리스트 (컬럼 키: code, codeName, typeName, typeVal, name)
    """
    if not excel_path.exists():
        raise FileNotFoundError(f"엑셀 파일을 찾을 수 없습니다: {excel_path}")
        
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    
    if not rows:
        return []
        
    header = rows[0]
    # 헤더 이름과 필드키 매핑
    field_keys = ["code", "codeName", "typeName", "typeVal", "name"]
    
    existing_data = []
    for row in rows[1:]:
        if not any(row):
            continue
        row_dict = {}
        for idx, key in enumerate(field_keys):
            val = str(row[idx]).strip() if idx < len(row) and row[idx] is not None else ""
            row_dict[key] = val
        existing_data.append(row_dict)
        
    return existing_data

def compare_code_data(existing_data: List[Dict[str, str]], latest_data: List[Dict[str, str]]) -> Dict[str, Any]:
    """
    기존 엑셀 데이터와 최신 API 데이터를 비교하여 변경 사항(신규, 수정, 삭제)을 추출하는 함수.
    
    Args:
        existing_data (List[Dict[str, str]]): 기존 엑셀 특성코드 리스트
        latest_data (List[Dict[str, str]]): 최신 API 특성코드 리스트
        
    Returns:
        Dict[str, Any]: 추가, 수정, 삭제, 유지 건수 및 변경 상세 내용
    """
    # 튜플 키 (code, typeVal, typeName, name, codeName) 복합키 또는 튜플 자체로 유일성 검사
    # 특성코드의 복합 식별자: (code, typeVal)
    # 다만 동일 (code, typeVal)에 여러건이 있을 수도 있으므로 (code, typeName, typeVal, name, codeName) 전체 매칭 및 복합키 매칭 진행
    
    def make_key(d: Dict[str, str]) -> Tuple[str, str, str, str, str]:
        return (
            str(d.get("code", "")).strip(),
            str(d.get("codeName", "")).strip(),
            str(d.get("typeName", "")).strip(),
            str(d.get("typeVal", "")).strip(),
            str(d.get("name", "")).strip()
        )

    def make_primary_key(d: Dict[str, str]) -> Tuple[str, str]:
        return (
            str(d.get("code", "")).strip(),
            str(d.get("typeVal", "")).strip()
        )

    # 1. 완전 일치 여부 비교를 위한 딕셔너리/집합 생성
    existing_full_set = set(make_key(item) for item in existing_data)
    latest_full_set = set(make_key(item) for item in latest_data)
    
    # 2. 식별키 (code, typeVal) 기준 맵 생성
    existing_pk_map: Dict[Tuple[str, str], List[Dict[str, str]]] = {}
    for item in existing_data:
        pk = make_primary_key(item)
        existing_pk_map.setdefault(pk, []).append(item)
        
    latest_pk_map: Dict[Tuple[str, str], List[Dict[str, str]]] = {}
    for item in latest_data:
        pk = make_primary_key(item)
        latest_pk_map.setdefault(pk, []).append(item)

    added_items = []
    modified_items = []
    removed_items = []
    unchanged_count = 0

    # 최신 데이터 기준 추가 및 수정 검사
    for item in latest_data:
        full_key = make_key(item)
        pk = make_primary_key(item)
        
        if full_key in existing_full_set:
            unchanged_count += 1
        else:
            # 완전 일치는 아니지만 PK가 기존에 존재하는 경우 -> 수정
            if pk in existing_pk_map:
                old_item = existing_pk_map[pk][0]
                modified_items.append({
                    "pk": pk,
                    "old": old_item,
                    "new": item
                })
            else:
                # PK도 없으면 -> 신규 추가
                added_items.append(item)

    # 기존 데이터 기준 삭제 검사
    for item in existing_data:
        full_key = make_key(item)
        pk = make_primary_key(item)
        if full_key not in latest_full_set and pk not in latest_pk_map:
            removed_items.append(item)

    return {
        "existing_total": len(existing_data),
        "latest_total": len(latest_data),
        "unchanged_count": unchanged_count,
        "added_count": len(added_items),
        "modified_count": len(modified_items),
        "removed_count": len(removed_items),
        "added_items": added_items,
        "modified_items": modified_items,
        "removed_items": removed_items
    }

def main():
    """
    메인 실행 함수: 엑셀 파일과 최신 API 응답을 비교 분석하고 결과를 출력합니다.
    """
    # Windows 콘솔 인코딩 대응
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    print(f"1. 기존 엑셀 파일 읽기: {EXCEL_PATH}")
    existing_data = read_existing_excel(EXCEL_PATH)
    print(f"   -> 기존 엑셀 데이터: {len(existing_data):,}건")
    
    print("2. 최신 특성코드 API(getCodeList) 호출 중...")
    latest_data = fetch_latest_code_list()
    print(f"   -> 최신 API 데이터: {len(latest_data):,}건")
    
    print("3. 특성코드 업데이트(변경 사항) 검사 진행 중...")
    result = compare_code_data(existing_data, latest_data)
    
    print("\n================ [특성코드 업데이트 검사 결과 리포트] ================")
    print(f" - 기존 엑셀 데이터 건수 : {result['existing_total']:,} 건")
    print(f" - 최신 API 데이터 건수  : {result['latest_total']:,} 건")
    print(f" - 변경 없는 데이터 건수 : {result['unchanged_count']:,} 건")
    print(f" - 신규 추가된 특성코드  : {result['added_count']:,} 건")
    print(f" - 내용 변경된 특성코드  : {result['modified_count']:,} 건")
    print(f" - 삭제된 특성코드       : {result['removed_count']:,} 건")
    print("======================================================================\n")

    if result["added_count"] > 0:
        print(f"--- [신규 추가 항목 ({result['added_count']}건 전체)] ---")
        for idx, item in enumerate(result["added_items"], 1):
            print(f"  {idx:2d}. 사양(code): {item.get('code')}, 사양명(codeName): {item.get('codeName')}, 특성명(typeName): {item.get('typeName')}, 특성값(typeVal): {item.get('typeVal')}, 이름(name): {item.get('name')}")
        print()

    if result["modified_count"] > 0:
        print(f"--- [내용 변경 항목 예시 (최대 10건)] ---")
        for m in result["modified_items"][:10]:
            pk_str = f"code={m['pk'][0]}, typeVal={m['pk'][1]}"
            print(f"  * 변경 대상 [{pk_str}]")
            print(f"    - 기존: 사양명={m['old'].get('codeName')}, 특성명={m['old'].get('typeName')}, 이름={m['old'].get('name')}")
            print(f"    + 변경: 사양명={m['new'].get('codeName')}, 특성명={m['new'].get('typeName')}, 이름={m['new'].get('name')}")
        print()

    if result["removed_count"] > 0:
        print(f"--- [삭제된 특성코드 상위 20건 예시 (전체 {result['removed_count']}건 중)] ---")
        for idx, item in enumerate(result["removed_items"][:20], 1):
            print(f"  {idx:2d}. 사양(code): {item.get('code')}, 특성값(typeVal): {item.get('typeVal')}, 특성명(typeName): {item.get('typeName')}, 이름(name): {item.get('name')}")
        print()

    if result["added_count"] == 0 and result["modified_count"] == 0 and result["removed_count"] == 0:
        print(">> 검사 결과: 기존 엑셀 파일과 최신 API 데이터가 100% 일치하며 업데이트된 사항이 없습니다.")

if __name__ == "__main__":
    main()
