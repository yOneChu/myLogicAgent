#!/usr/bin/env python3
"""
executeQuery API 결과(JSON)를 Excel(xlsx) 파일로 저장하는 스크립트.

사용 예시:
  python query_to_excel.py --sql "SELECT * FROM HDEL_DEFAULT.SOME_TABLE" --output result.xlsx
  python query_to_excel.py --sql-file query.sql --output result.xlsx
  py query_to_excel.py --sql-file query.sql --output result.xlsx

결과 파일은 작업 공간 최상위의 output_excel 폴더에 생성됩니다.

사내 인증서 문제로 SSL 오류가 발생할 때만:
  python query_to_excel.py --sql-file query.sql --output result.xlsx --insecure
"""

import argparse
import json
import ssl
import sys
from collections import OrderedDict
from pathlib import Path
from typing import Any
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import Workbook
from openpyxl.styles import PatternFill


API_URL = "https://vault-in.hdel.co.kr:8070/api/executeQuery"
API_KEY = "subae"

# scripts/db_query/query_to_excel.py에서 실행되므로
# __file__이 scripts/db_query/에 위치함. parent.parent.parent를 통해 워크스페이스 최상위로 이동.
OUTPUT_DIR = Path(__file__).resolve().parent.parent.parent / "output_excel"

HEADER_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


HEADER_KEYS = ("headers", "header", "columns", "cols", "fields")
DATA_KEYS = ("data", "rows", "resultList", "list", "items", "results")


def read_sql(args: argparse.Namespace) -> str:
    if args.sql:
        return args.sql.strip()

    if args.sql_file:
        with open(args.sql_file, "r", encoding="utf-8") as f:
            return f.read().strip()

    raise ValueError("--sql 또는 --sql-file 중 하나는 반드시 입력해야 합니다.")


def request_json(sql: str, insecure: bool = False) -> Any:
    query_string = urlencode({"key": API_KEY, "sql": sql})
    url = f"{API_URL}?{query_string}"

    req = Request(
        url,
        method="GET",
        headers={
            "Accept": "application/json",
            "User-Agent": "query-to-excel/1.0",
        },
    )

    context = None
    if insecure:
        context = ssl._create_unverified_context()

    with urlopen(req, timeout=120, context=context) as response:
        raw_body = response.read()

    try:
        body = raw_body.decode("utf-8")
        if "\ufffd" in body:
            body = raw_body.decode("cp949")
    except (UnicodeDecodeError, UnicodeEncodeError):
        body = raw_body.decode("cp949", errors="replace")

    try:
        return json.loads(body)
    except json.JSONDecodeError as exc:
        raise ValueError(f"API 응답이 JSON 형식이 아닙니다: {body[:500]}") from exc


def normalize_header_name(header: Any) -> str:
    if isinstance(header, dict):
        for key in ("name", "field", "column", "columnName", "label", "title", "id"):
            if key in header and header[key] is not None:
                return str(header[key])
        return str(header)

    return str(header)


def normalize_headers(headers: Any) -> list[str]:
    if not headers:
        return []

    return [normalize_header_name(header) for header in headers]


def find_table_payload(payload: Any) -> tuple[list[str], list[Any]]:
    """
    API마다 JSON 래핑 구조가 다를 수 있어 대표 패턴을 순서대로 찾는다.

    지원 예:
      {"headers": [...], "data": [...]}
      {"columns": [...], "rows": [...]}
      {"result": {"headers": [...], "data": [...]}}
      [{"COL1": "A", "COL2": "B"}]
    """
    if isinstance(payload, list):
        return [], payload

    if not isinstance(payload, dict):
        raise ValueError("지원하지 않는 JSON 응답 구조입니다.")

    headers = []
    rows = None

    for key in HEADER_KEYS:
        if key in payload:
            headers = normalize_headers(payload[key])
            break

    for key in DATA_KEYS:
        if key in payload:
            rows = payload[key]
            break

    if isinstance(rows, list):
        return headers, rows

    for value in payload.values():
        if isinstance(value, (dict, list)):
            try:
                return find_table_payload(value)
            except ValueError:
                pass

    raise ValueError("Excel로 변환할 데이터 목록을 찾지 못했습니다.")


def infer_headers_from_dict_rows(rows: list[dict[str, Any]]) -> list[str]:
    headers: OrderedDict[str, None] = OrderedDict()

    for row in rows:
        for key in row.keys():
            headers.setdefault(str(key), None)

    return list(headers.keys())


def rows_to_records(headers: list[str], rows: list[Any]) -> tuple[list[str], list[dict[str, Any]]]:
    if not rows:
        return headers, []

    first_row = rows[0]

    if isinstance(first_row, dict):
        if not headers:
            headers = infer_headers_from_dict_rows(rows)

        records = []
        for row in rows:
            records.append({header: row.get(header, "") for header in headers})

        return headers, records

    if isinstance(first_row, (list, tuple)):
        if not headers:
            max_len = max(len(row) for row in rows)
            headers = [f"COL{i}" for i in range(1, max_len + 1)]

        records = []
        for row in rows:
            record = {}
            for index, header in enumerate(headers):
                record[header] = row[index] if index < len(row) else ""
            records.append(record)

        return headers, records

    if headers:
        return headers, [{headers[0]: row} for row in rows]

    return ["VALUE"], [{"VALUE": row} for row in rows]


def to_cell_value(value: Any) -> Any:
    if value is None:
        return ""

    if isinstance(value, (str, int, float, bool)):
        return value

    return str(value)


def write_excel(path: Path, headers: list[str], records: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Result"

    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL

    for record in records:
        sheet.append([to_cell_value(record.get(header)) for header in headers])

    workbook.save(path)


def resolve_output_path(output: str) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    path = OUTPUT_DIR / Path(output).name
    if path.suffix.lower() != ".xlsx":
        path = path.with_suffix(".xlsx")

    return path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="executeQuery API를 GET 방식으로 호출하고 JSON 결과를 Excel(xlsx)로 저장합니다."
    )
    parser.add_argument("--sql", help="실행할 SQL 문장")
    parser.add_argument("--sql-file", help="SQL이 저장된 파일 경로")
    parser.add_argument("--output", default="result.xlsx", help="저장할 Excel 파일명 (output_excel 폴더에 생성)")
    parser.add_argument(
        "--insecure",
        action="store_true",
        help="SSL 인증서 검증을 생략합니다. 사내 인증서 오류가 날 때만 사용하세요.",
    )
    return parser.parse_args()


def main() -> int:
    try:
        args = parse_args()
        sql = read_sql(args)

        payload = request_json(sql, insecure=args.insecure)
        headers, rows = find_table_payload(payload)
        headers, records = rows_to_records(headers, rows)

        if not headers:
            raise ValueError("Excel 헤더를 만들 수 없습니다.")

        output_path = resolve_output_path(args.output)
        write_excel(output_path, headers, records)
        print(f"Excel 저장 완료: {output_path} / {len(records)}건")
        return 0

    except Exception as exc:
        print(f"오류: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
