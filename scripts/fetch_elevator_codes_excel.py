#!/usr/bin/env python3
"""
엘리베이터 특성코드 API(getCodeList)를 호출하여 28,000여 건의 대용량 데이터를 수집하고
엑셀(xlsx) 파일로 고속 변환하여 master_data 폴더에 저장하는 스크립트.
"""

import json
import ssl
import sys
import urllib.request
import urllib.parse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# API URL 및 접근 키 설정
API_URL = "https://vault-in.hdel.co.kr:8070/api/getCodeList"
API_KEY = "subae"

# 워크스페이스 최상위 경로 및 출력 대상 master_data 디렉터리 지정
WORKSPACE_ROOT = Path(__file__).resolve().parent.parent
MASTER_DATA_DIR = WORKSPACE_ROOT / "master_data"


def fetch_elevator_code_list() -> list:
    """
    엘리베이터 특성코드 API를 GET 방식으로 호출하여 전체 특성코드 리스트 데이터를 반환합니다.
    
    Returns:
        list: API 응답으로 받은 특성코드 데이터 dictionary 리스트
    """
    query_string = urllib.parse.urlencode({"key": API_KEY})
    url = f"{API_URL}?{query_string}"
    
    req = urllib.request.Request(
        url,
        method="GET",
        headers={
            "Accept": "application/json",
            "User-Agent": "elevator-code-fetcher/1.0"
        }
    )
    
    # 사내 SSL 인증서 검증 생략 옵션 적용
    ctx = ssl._create_unverified_context()
    
    with urllib.request.urlopen(req, timeout=120, context=ctx) as response:
        raw_bytes = response.read()
        
    # UTF-8 및 CP949 인코딩 순차 디코딩
    try:
        text = raw_bytes.decode("utf-8")
    except UnicodeDecodeError:
        text = raw_bytes.decode("cp949", errors="replace")
        
    data = json.loads(text)
    
    if isinstance(data, list):
        return data
    elif isinstance(data, dict):
        for key in ("data", "result", "items", "rows", "list"):
            if key in data and isinstance(data[key], list):
                return data[key]
    
    raise ValueError("API 응답 데이터 형식을 해석할 수 없습니다.")


def export_to_excel_fast(data: list, output_filename: str = "엘리베이터_특성코드.xlsx") -> Path:
    """
    대용량 특성코드 리스트를 고속으로 엑셀 파일로 작성하고 master_data 디렉터리에 저장합니다.
    
    Args:
        data (list): 수집된 특성코드 JSON 리스트
        output_filename (str): 생성할 엑셀 파일명
        
    Returns:
        Path: 저장된 엑셀 파일의 절대 경로
    """
    # master_data 폴더가 없을 경우 자동 생성
    MASTER_DATA_DIR.mkdir(parents=True, exist_ok=True)
    output_path = MASTER_DATA_DIR / output_filename

    wb = Workbook()
    ws = wb.active
    ws.title = "엘리베이터_특성코드"
    
    # 헤더 명칭 및 필드키 정의
    headers = ["사양 (code)", "사양명 (codeName)", "특성명 (typeName)", "특성값 (typeVal)", "이름 (name)"]
    field_keys = ["code", "codeName", "typeName", "typeVal", "name"]
    
    # 1행 헤더 추가
    ws.append(headers)
    
    # 헤더 스타일링 설정
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    # 고속 데이터 삽입 (셀별 개별 스타일 루프를 배제하여 속도 극대화)
    for row_data in data:
        ws.append([row_data.get(k, "") for k in field_keys])

    # 컬럼 너비 적절한 고정값 설정 (대용량 연산 방지)
    column_widths = {"A": 18, "B": 28, "C": 28, "D": 28, "E": 28}
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 엑셀 파일 저장
    wb.save(output_path)
    return output_path


def main():
    """
    스크립트 메인 실행 함수: 특성코드 데이터를 수집하여 master_data 폴더에 엑셀로 저장합니다.
    """
    print("엘리베이터 특성코드 API 데이터를 요청합니다...")
    data = fetch_elevator_code_list()
    print(f"API 호출 완료: 총 {len(data):,}건 수집됨.")
    
    excel_path = export_to_excel_fast(data, "엘리베이터_특성코드.xlsx")
    print(f"엑셀 저장 성공: {excel_path}")


if __name__ == "__main__":
    main()
