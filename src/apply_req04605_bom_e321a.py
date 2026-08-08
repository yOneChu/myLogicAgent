#!/usr/bin/env python3
"""
전산화 요청 04605 처리를 위한 BOM_E321A PID 로직 반영 스크립트입니다.
기존 BOM_E321A (v8) 로직 데이터에 신규 정합성 체크 행(NO 10)을 추가하여
output_csv 및 output_excel 폴더에 결과물 파일(CSV, Excel)을 생성합니다.
"""

import csv
import sys
from pathlib import Path
from typing import Dict, List, Any, Tuple

from openpyxl import Workbook
from openpyxl.styles import PatternFill

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

# 프로젝트 루트 및 경로 설정
ROOT_DIR = Path(__file__).resolve().parent.parent
INPUT_CSV = ROOT_DIR / "output_csv" / "bom_e321a_full.csv"
OUTPUT_CSV = ROOT_DIR / "output_csv" / "BOM_E321A_req04605.csv"
OUTPUT_EXCEL = ROOT_DIR / "output_excel" / "BOM_E321A_req04605.xlsx"

HEADER_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


def load_original_bom_e321a(csv_path: Path) -> Tuple[List[str], List[Dict[str, str]]]:
    """
    기존 BOM_E321A의 CSV 파일 데이터를 읽어 헤더 목록과 행 데이터 리스트를 반환하는 함수입니다.
    
    :param csv_path: 원본 CSV 파일 경로
    :return: (headers, rows) 튜플
    """
    rows = []
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames if reader.fieldnames else []
        for row in reader:
            rows.append(dict(row))
    return list(headers), rows


def build_new_logic_row(existing_headers: List[str]) -> Dict[str, str]:
    """
    전산화 요청 04605에 정의된 조건 및 산출 결과를 바탕으로 신규 정합성 로직 행(NO 10)을 생성하는 함수입니다.
    
    :param existing_headers: CSV 헤더 목록
    :return: 신규 생성된 행 데이터 딕셔너리
    """
    new_row = {h: "-" for h in existing_headers}
    
    # 기본 정보 설정
    new_row["PID"] = "BOM_E321A"
    new_row["VERSION"] = "8"
    new_row["NO"] = "10"
    new_row["ADDR"] = "-"
    new_row["GOTO"] = "-"
    
    # 조건 (SPEC / CON) 설정
    # SPEC1: 주석 정보 특성코드 (E321A_CMT / EL_ACMT)
    new_row["SPEC1"] = "E321A_CMT"
    new_row["CON1"] = ",?V4612?,?V04612?,?V004612?,?V0004612?"
    
    # SPEC2: BG 비표준 값 판단 조건 (BG != CA + 100)
    new_row["SPEC2"] = "EL_ECBG"
    new_row["CON2"] = "![$ {EL_ECCA} + 100 $]"
    
    # 산출 결과 (KEY / VAL) 설정
    # KEY1: 신규 주석 문구 L_CMT 생성
    new_row["KEY1"] = "L_CMT"
    new_row["VAL1"] = "V0004612 적용 시 CEILING ASSY(E321A) 및 ISOLATION BRKT(E321A16) 주석에 KE =[$ {K4} $], KF=[$ {K4} + 20 $]로 변경 적용돼었는지 확인할 것. (BG 비표준 값 반영)"
    
    # KEY2: 기존 주석에 누적 이어나감
    new_row["KEY2"] = "CMT"
    new_row["VAL2"] = "{CMT}{L_CMT}"
    
    # KEY3: 오류 등급 WARNING
    new_row["KEY3"] = "ERR_LV"
    new_row["VAL3"] = "WARNING"
    
    # REMARKS: 적용 사유 및 안건 메모
    new_row["REMARKS"] = "V0004612 도면 적용 시 BG 비표준 값 정합성 안내 주석(요청 04605)"
    
    return new_row


def write_result_csv(csv_path: Path, headers: List[str], rows: List[Dict[str, str]]) -> None:
    """
    최종 로직 데이터 리스트를 UTF-8-SIG 인코딩의 CSV 파일로 저장하는 함수입니다.
    
    :param csv_path: 저장할 CSV 파일 경로
    :param headers: 헤더 컬럼 목록
    :param rows: 행 데이터 리스트
    """
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def write_result_excel(excel_path: Path, headers: List[str], rows: List[Dict[str, str]]) -> None:
    """
    최종 로직 데이터 리스트를 Excel (.xlsx) 파일로 저장하는 함수입니다.
    
    :param excel_path: 저장할 Excel 파일 경로
    :param headers: 헤더 컬럼 목록
    :param rows: 행 데이터 리스트
    """
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM_E321A_req04605"
    
    # 헤더 행 추가
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        
    # 데이터 행 추가
    for row in rows:
        ws.append([row.get(h, "-") for h in headers])
        
    wb.save(excel_path)


def main():
    """
    메인 실행 함수로, 원본 로직 읽기 -> 신규 행 추가 -> CSV 및 Excel 저장을 수행합니다.
    """
    print(f"[1/3] 기존 BOM_E321A 데이터를 읽어오는 중입니다: {INPUT_CSV}")
    headers, rows = load_original_bom_e321a(INPUT_CSV)
    
    print(f"[2/3] 전산화 요청 04605에 따른 신규 정합성 로직 행(NO 10) 추가 중...")
    new_row = build_new_logic_row(headers)
    rows.append(new_row)
    
    print(f"[3/3] 결과물 파일 생성 중...")
    write_result_csv(OUTPUT_CSV, headers, rows)
    print(f" -> CSV 저장 완료: {OUTPUT_CSV}")
    
    write_result_excel(OUTPUT_EXCEL, headers, rows)
    print(f" -> Excel 저장 완료: {OUTPUT_EXCEL}")
    
    print("\n=== 반영 결과 데이터 요약 ===")
    print(f"총 행 수: {len(rows)}행")
    print(f"신규 추가된 NO 10 행 내역:")
    print(f"  - ADDR/GOTO: {new_row['ADDR']} / {new_row['GOTO']}")
    print(f"  - SPEC1 / CON1: {new_row['SPEC1']} = {new_row['CON1']}")
    print(f"  - SPEC2 / CON2: {new_row['SPEC2']} = {new_row['CON2']}")
    print(f"  - KEY1 / VAL1: {new_row['KEY1']} = {new_row['VAL1']}")
    print(f"  - KEY2 / VAL2: {new_row['KEY2']} = {new_row['VAL2']}")
    print(f"  - KEY3 / VAL3: {new_row['KEY3']} = {new_row['VAL3']}")
    print(f"  - REMARKS: {new_row['REMARKS']}")


if __name__ == "__main__":
    main()
